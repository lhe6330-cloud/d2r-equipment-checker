import sys
import os

# 模拟打包后的环境
if hasattr(sys, '_MEIPASS'):
    print(f"运行在打包环境中: _MEIPASS = {sys._MEIPASS}")
    exe_dir = os.path.dirname(sys.executable)
    print(f"EXE 目录：{exe_dir}")
else:
    print("运行在 Python 环境中")
    exe_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"脚本目录：{exe_dir}")

# 检查 keywords.xlsx 是否存在
keywords_path = os.path.join(exe_dir, 'keywords.xlsx')
print(f"\n检查 keywords.xlsx: {keywords_path}")
print(f"文件存在：{os.path.exists(keywords_path)}")

# 尝试加载
if os.path.exists(keywords_path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(keywords_path, data_only=True)
        ws = wb.active
        print(f"\n成功加载 Excel!")
        print(f"总行数：{ws.max_row}")
        
        # 显示前 10 行
        print("\n前 10 行关键词:")
        for row in range(1, min(11, ws.max_row + 1)):
            en = ws.cell(row=row, column=2).value
            cn = ws.cell(row=row, column=3).value
            print(f"  {en} -> {cn}")
    except Exception as e:
        print(f"\n加载失败：{e}")
        import traceback
        traceback.print_exc()
else:
    print(f"\n错误：keywords.xlsx 不存在于 {exe_dir}")
    print(f"当前目录内容：{os.listdir(exe_dir)}")
