"""
Configuration constants and mappings for DD373 D2R Equipment Checker.
"""
from dataclasses import dataclass, field
from typing import Dict, List, Optional
from enum import Enum
import sys
import os
from pathlib import Path


# =============================================================================
# File Paths - Handle both dev and bundled exe environments
# =============================================================================
def get_data_dir():
    """Get the directory for storing user data files."""
    if hasattr(sys, '_MEIPASS'):
        # Running as bundled exe - store data next to the exe
        return os.path.dirname(sys.executable)
    # Running in normal Python environment
    return os.path.dirname(os.path.abspath(__file__)) if '__file__' in dir() else '.'

DATA_DIR = get_data_dir()
OUTPUT_FILE = os.path.join(DATA_DIR, "output.txt")
ORDERLIST_FILE = os.path.join(DATA_DIR, "orderlist.json")
UI_FILE = "d2rcheck.ui"


# =============================================================================
# HTTP Configuration
# =============================================================================
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/91.0.4472.124 Safari/537.36"
)
REQUEST_TIMEOUT = 15
PAGE_DELAY_SECONDS = 5

# =============================================================================
# Currency Configuration
# =============================================================================
CNY_TO_USD_RATE = 0.14  # 1 CNY ≈ 0.14 USD (adjust as needed)


def format_cny(price: str) -> str:
    """Format price as CNY."""
    try:
        clean_price = price.replace('¥', '').replace(',', '').strip()
        cny_value = float(clean_price)
        return f"¥{cny_value:.2f}"
    except (ValueError, AttributeError):
        return f"¥{price}"


def format_usd(price: str) -> str:
    """Convert CNY price to USD."""
    try:
        clean_price = price.replace('¥', '').replace(',', '').strip()
        cny_value = float(clean_price)
        usd_value = cny_value * CNY_TO_USD_RATE
        return f"${usd_value:.2f}"
    except (ValueError, AttributeError):
        return "N/A"


def format_price(cny_price: str) -> str:
    """
    Format price to show both CNY and USD.
    
    Args:
        cny_price: Price string in CNY (may contain ¥ symbol)
    
    Returns:
        Formatted string like "¥580 / $81"
    """
    try:
        # Clean the price string
        clean_price = cny_price.replace('¥', '').replace(',', '').strip()
        cny_value = float(clean_price)
        usd_value = cny_value * CNY_TO_USD_RATE
        return f"¥{cny_value:.0f} / ${usd_value:.2f}"
    except (ValueError, AttributeError):
        return f"¥{cny_price}"


# =============================================================================
# Task Status Icons
# =============================================================================
class TaskStatus(Enum):
    WAITING = "◯ waiting"
    CHECKING = "▶ checking"
    PAUSED = "⏸ pause"
    FINISHED = "✓ finished"
    STOPPED = "✗ stop"
    SKIPPED = "user skipped"


# =============================================================================
# Special Attributes (existence-based, no numeric value required)
# =============================================================================
SPECIAL_ATTRIBUTES: Dict[str, List[str]] = {
    'socket': ['socket', '孔'],
    'Ethereal': ['Ethereal', 'Eth', '无形'],
    'Eth and Repairs durability persecond': ['Eth and Repairs durability persecond', 'erep', '无形自回'],
    'Eth and Replenishes Quantity': ['Eth and Replenishes Quantity', 'erep', '无形自回'],
    'sigil:lethargy': ['sigil:lethargy', '迟缓'],
    'miasma bolt': ['miasma bolt', '毒雾弹'],
}


# =============================================================================
# Attribute Mappings (numeric value-based)
# Generated from Filter Word Mapping Table.xlsx
# =============================================================================
ATTRIBUTE_MAP: Dict[str, List[str]] = {
    'Faster cast rate': ['Faster cast rate', 'fcr', '施法'],
    'Faster Hit Recovery': ['Faster Hit Recovery', 'Fhr', '打击恢复'],
    'Faster Run/Walk': ['Faster Run/Walk', 'frw', '跑'],
    'strength': ['strength', 'str', '力量'],
    'dexterity': ['dexterity', 'dex', '敏捷'],
    'life': ['life', '生命'],
    'vitality': ['vitality', 'vit', '体能'],
    'energy': ['energy', 'ene', '精力'],
    'all res': ['all res', 'allres', '所有抗性'],
    'fire resistance': ['fire resistance', 'fr', '火抗'],
    'cold resistance': ['cold resistance', 'cr', '冰抗'],
    'lightning resistance': ['lightning resistance', 'lr', '电抗'],
    'poison resistance': ['poison resistance', 'pr', '毒抗'],
    'magic find': ['magic find', 'mf', '寻宝'],
    'extro gold': ['extro gold', 'eg', '打钱'],
    'mana': ['mana', '法力'],
    'defense': ['defense', 'def', '防御'],
    'Increased attack speed': ['Increased attack speed', 'ias', '攻击速度'],
    'life steal': ['life steal', 'll', '吸血'],
    'mana steal': ['mana steal', 'lm', '吸蓝'],
    'Enhanced Damage/': ['Enhanced Damage/', 'ed', '伤害'],
    'Enhanced Defense': ['Enhanced Defense', 'ed', '防御'],
    'attack rating': ['attack rating', 'ar', '准确率'],
    'socket': ['socket', '孔'],
    'Chance Of Crushing Blow': ['Chance Of Crushing Blow', 'cb', '压碎打击'],
    'Replenish Life': ['Replenish Life', 'replife', '生命恢复'],
    'Chance Of Deadly Strike': ['Chance Of Deadly Strike', 'ds', '致命打击'],
    'To Attack Rating(Based On Character Level)': ['To Attack Rating(Based On Character Level)', 'fools', '幻影'],
    'Bonus To Attack Rating (Based On Character Level)': ['Bonus To Attack Rating (Based On Character Level)', 'visio', '角色等级'],
    'Max dmg': ['Max dmg', 'max', '最大伤害'],
    'Min dmg': ['Min dmg', 'min', '最小伤害'],
    'Poison Length Reduced by': ['Poison Length Reduced by', 'plr', '中毒时间减少'],
    'to Poison and Bone Skills(Necromancer Only)': ['to Poison and Bone Skills(Necromancer Only)', 'pnb', '毒骨'],
    'Ethereal': ['Ethereal', 'Eth', '无形'],
    'Eth and Repairs durability persecond': ['Eth and Repairs durability persecond', 'erep', '无形自回'],
    'Eth and Replenishes Quantity': ['Eth and Replenishes Quantity', 'erep', '无形自回'],
    'teleport': ['teleport', 'tp', '传送'],
    'Increased Chance of Blocking': ['Increased Chance of Blocking', 'icb', '增加格挡'],
    'Mind blast': ['Mind blast', 'mb', '心灵爆震'],
    'Lightning Sentry': ['Lightning Sentry', 'ls', '雷光哨卫'],
    'Wake of Fire': ['Wake of Fire', 'wof', '火焰复苏'],
    'Blade Fury': ['Blade Fury', 'bf', '刀刃之怒'],
    'Fire Blast': ['Fire Blast', 'FB', '火焰爆'],
    'Fire ball': ['Fire ball', 'FB', '火球'],
    'Shadow Master': ['Shadow Master', 'sm', '影子大师'],
    'Venom': ['Venom', '毒牙'],
    'Weapon Block': ['Weapon Block', 'Wb', '武器格挡'],
    'Shadow Warrior': ['Shadow Warrior', 'sw', '影子战士'],
    'Cloak of Shadows': ['Cloak of Shadows', 'cos', '魔影斗篷'],
    'Fade': ['Fade', '能量消解'],
    'Claws of Thunder': ['Claws of Thunder', 'cot', '雷电爪'],
    'Blades of Ice': ['Blades of Ice', 'boi', '寒冰刃'],
    'Phoenix Strike': ['Phoenix Strike', 'ps', '凤凰击'],
    'Dragon Tail': ['Dragon Tail', 'dt', '神龙摆尾'],
    'assassin': ['assassin', 'sin', '刺客'],
    'Shadow skill': ['Shadow skill', 'Shadow', '暗影'],
    'Trap skill': ['Trap skill', 'trap', '陷阱'],
    'Blade sheild': ['Blade sheild', 'bs', '刃之盾'],
    'Amplify Damage': ['Amplify Damage', 'AMP', '伤害加深'],
    'Terror': ['Terror', '恐惧'],
    'Confuse': ['Confuse', '迷乱'],
    'Life Tap': ['Life Tap', 'LT', '偷取生命'],
    'Attract': ['Attract', '吸引'],
    'Decrepify': ['Decrepify', 'Decrep', '衰老'],
    'Lower Resist': ['Lower Resist', 'LR', '降低抵抗'],
    'Bone Spear': ['Bone Spear', 'BS', '骨矛'],
    'Bone Spirit': ['Bone Spirit', 'BSpirit', '骨魂'],
    'Poison Nova': ['Poison Nova', 'PN', '毒新星'],
    'Bone Prison': ['Bone Prison', 'BP', '骨牢'],
    'Blessed Hammer': ['Blessed Hammer', 'BH', '祝福之锤'],
    'Golem Mastery': ['Golem Mastery', 'GM', '石魔支配'],
    'Fire Golem': ['Fire Golem', 'FG', '火焰石魔'],
    'Blood Golem': ['Blood Golem', 'Bg', '鲜血石魔'],
    'Clay Golem': ['Clay Golem', 'cg', '黏土石魔'],
    'Bone Wall': ['Bone Wall', 'BWall', '骨墙'],
    'Charge': ['Charge', '冲锋'],
    'Concentration': ['Concentration', 'Conc', '专注'],
    'Conviction': ['Conviction', 'Conv', '信念'],
    'Fist of the Heavens': ['Fist of the Heavens', 'Foh', '天堂之拳'],
    'Combat': ['Combat', '战斗技能'],
    'Offensive': ['Offensive', '攻击灵气'],
    'Shiver Armor': ['Shiver Armor', 'SA', '碎冰甲'],
    'Chilling Armor': ['Chilling Armor', 'ca', '寒冰装甲'],
    'Frozen Orb': ['Frozen Orb', 'fo', '冰封球'],
    'Cold Mastery': ['Cold Mastery', 'CM', '冰冷支配'],
    'Blizzard': ['Blizzard', 'blizz', '暴风雪'],
    'Nova': ['Nova', '新星'],
    'Lightning': ['Lightning', '闪电'],
    'Chain Lightning': ['Chain Lightning', 'CL', '连锁闪电'],
    'Thunder Storm': ['Thunder Storm', 'ts', '雷云风暴'],
    'Energy Shield': ['Energy Shield', 'es', '能量护盾'],
    'Lightning Mastery': ['Lightning Mastery', 'lm', '闪电支配'],
    'Fire Wall': ['Fire Wall', 'fw', '火墙'],
    'Meteor': ['Meteor', '陨石'],
    'Hydra': ['Hydra', '九头蛇'],
    'Enchant': ['Enchant', '强化'],
    'Fire Mastery': ['Fire Mastery', 'fm', '火焰支配'],
    'Frost Nova': ['Frost Nova', 'FN', '霜之新星'],
    'Fissure': ['Fissure', '火山爆'],
    'Volcano': ['Volcano', 'Volc', '火山'],
    'Cyclone Armor': ['Cyclone Armor', 'Cyclone', '飓风装甲'],
    'Tornado': ['Tornado', 'nado', '龙卷风'],
    'Hurricane': ['Hurricane', '暴风'],
    'Maul': ['Maul', '撞槌'],
    'Oak Sage': ['Oak Sage', 'oak', '橡木智者'],
    'Fury': ['Fury', '狂怒'],
    'Heart of Wolverine': ['Heart of Wolverine', 'HoW', '狼獾之心'],
    'Firestorm': ['Firestorm', 'FS', '火风暴'],
    'ring of fire': ['ring of fire', '火焰环'],
    'flame wave': ['flame wave', '火焰波'],
    'apocalypse': ['apocalypse', '启示'],
    'sigil:lethargy': ['sigil:lethargy', '迟缓'],
    'sigil:rancor': ['sigil:rancor', '怨恨'],
    'sigil:death': ['sigil:death', '死亡'],
    'miasma bolt': ['miasma bolt', '毒雾弹'],
    'miasma chain': ['miasma chain', '毒雾链'],
    'enhanced entropy': ['enhanced entropy', '增强熵'],
    'abyss': ['abyss', '深渊'],
    'levitation mastery': ['levitation mastery', '漂浮支配'],
    'echoing strike': ['echoing strike', '回音击'],
    'blade warp': ['blade warp', '刃之扭曲'],
    'cleave': ['cleave', '劈砍'],
    'psychic ward': ['psychic ward', '心灵结界'],
    'eldritch blast': ['eldritch blast', '诡异爆'],
    'mirrored blades': ['mirrored blades', '镜像之刃'],
    'hex:bane': ['hex:bane', '诅咒'],
    'hex:purge': ['hex:purge', '净化'],
    'hex:siphon': ['hex:siphon', '吸取'],
    'demonic mastery': ['demonic mastery', '恶魔支配'],
    'consume': ['consume', '吞噬'],
    'engorge': ['engorge', '吞食'],
    'blood boil': ['blood boil', '血液沸腾'],
    'bind demon': ['bind demon', '束缚恶魔'],
}


# =============================================================================
# Gear Type URL Templates
# =============================================================================
@dataclass
class GearTypeUrls:
    """URLs for a specific gear type with 5 mode options."""
    all: str
    ladder: str
    warlock_ladder: str
    warlock_nonladder: str
    nonladder: str


GEAR_TYPE_URLS: Dict[str, GearTypeUrls] = {
    'weapon': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-915cc3-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-915cc3-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-915cc3-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-915cc3-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-915cc3-0-0-0-0-0-0-0-0.html'
    ),
    'hat': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-fg6vuh-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-fg6vuh-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-fg6vuh-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-fg6vuh-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-fg6vuh-0-0-0-0-0-0-0-0.html'
    ),
    'shield': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-krwbfs-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-krwbfs-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-krwbfs-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-krwbfs-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-krwbfs-0-0-0-0-0-0-0-0.html'
    ),
    'armor': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-39bt7g-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-39bt7g-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-39bt7g-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-39bt7g-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-39bt7g-0-0-0-0-0-0-0-0.html'
    ),
    'belt': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-f3q9tg-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-f3q9tg-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-f3q9tg-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-f3q9tg-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-f3q9tg-0-0-0-0-0-0-0-0.html'
    ),
    'glove': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-uhd8dt-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-uhd8dt-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-uhd8dt-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-uhd8dt-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-uhd8dt-0-0-0-0-0-0-0-0.html'
    ),
    'boots': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-3s1baw-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-3s1baw-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-3s1baw-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-3s1baw-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-3s1baw-0-0-0-0-0-0-0-0.html'
    ),
    'ring': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-hquct0-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-hquct0-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-hquct0-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-hquct0-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-hquct0-0-0-0-0-0-0-0-0.html'
    ),
    'amulet': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-v41x85-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-v41x85-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-v41x85-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-v41x85-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-v41x85-0-0-0-0-0-0-0-0.html'
    ),
    'charm': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-uv1bd8-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-uv1bd8-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-uv1bd8-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-uv1bd8-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-uv1bd8-0-0-0-0-0-0-0-0.html'
    ),
    'jewel': GearTypeUrls(
        all='https://www.dd373.com/s-1psrbm-u6w1hm-0-0-0-0-5tgw08-0tc73c-0-0-0-0-0-0-0-0.html',
        ladder='https://www.dd373.com/s-1psrbm-u6w1hm-7wme4a-0-0-0-5tgw08-0tc73c-0-0-0-0-0-0-0-0.html',
        warlock_ladder='https://www.dd373.com/s-1psrbm-u6w1hm-hx35xs-0-0-0-5tgw08-0tc73c-0-0-0-0-0-0-0-0.html',
        warlock_nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-6mu7b4-0-0-0-5tgw08-0tc73c-0-0-0-0-0-0-0-0.html',
        nonladder='https://www.dd373.com/s-1psrbm-u6w1hm-dkcdwk-0-0-0-5tgw08-0tc73c-0-0-0-0-0-0-0-0.html'
    ),
}


def get_gear_type_url(gear_type: str, mode: str) -> str:
    """
    Get the search URL for a given gear type and mode.
    
    Args:
        gear_type: Type of gear (weapon, hat, shield, etc.) - case insensitive
        mode: Mode string (all, ladder, warlock ladder, warlock nonladder, nonladder)
    
    Returns:
        The appropriate search URL
    """
    gear_type_lower = gear_type.lower()
    
    # Handle alternate names
    gear_type_aliases = {
        'sheild': 'shield',  # Common typo
        'helm': 'hat',
        'helmet': 'hat',
        'circlet': 'hat',
        'gloves': 'glove',
        'boot': 'boots',
    }
    gear_type_lower = gear_type_aliases.get(gear_type_lower, gear_type_lower)
    
    # Normalize mode string (handle variations in input)
    mode_key = mode.lower().replace(' ', '_').replace('-', '_').replace(')', '')
    
    # Map common variations to standard keys
    mode_mapping = {
        'all': 'all',
        'ladder': 'ladder',
        'warlock_ladder': 'warlock_ladder',
        'warlock_nonladder': 'warlock_nonladder',
        'nonladder': 'nonladder',
        'nl_only': 'nonladder',
        'l_only': 'ladder',
        'both': 'all',
        'nonladder_ladder': 'all',  # Legacy support
    }
    mode_key = mode_mapping.get(mode_key, mode_key)
    
    # Get URLs for gear type
    gear_urls = GEAR_TYPE_URLS.get(gear_type_lower)
    if gear_urls:
        return getattr(gear_urls, mode_key, gear_urls.nonladder)
    
    # Default to amulet nonladder
    return GEAR_TYPE_URLS['amulet'].nonladder


# =============================================================================
# Column Width Configuration
# =============================================================================
@dataclass
class TableColumnWidths:
    """Column widths for table widgets."""
    main_table: Dict[int, int] = field(default_factory=lambda: {
        0: 300,  # stats
        1: 80,   # price
        2: 150,  # link
        3: 120,  # add to order
    })
    order_list: Dict[int, int] = field(default_factory=lambda: {
        0: 300,  # stats
        1: 80,   # price
        2: 150,  # link
        3: 80,   # copy link
        4: 60,   # delete
    })
    task_list: Dict[int, int] = field(default_factory=lambda: {
        0: 200,  # name
        1: 100,  # status
        2: 60,   # delete
    })
    task_result: Dict[int, int] = field(default_factory=lambda: {
        0: 150,  # task name
        1: 250,  # stats
        2: 80,   # price
        3: 150,  # link
        4: 120,  # add to order
        5: 60,   # delete
    })


COLUMN_WIDTHS = TableColumnWidths()


# =============================================================================
# Keywords Configuration
# =============================================================================

def get_keywords_file_path() -> str:
    """Get the path to keywords.xlsx file."""
    if hasattr(sys, '_MEIPASS'):
        # Running as bundled exe
        return os.path.join(os.path.dirname(sys.executable), 'keywords.xlsx')
    # Running in normal Python environment
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'keywords.xlsx')


def load_keyword_mapping(excel_path: str = None) -> Dict[str, str]:
    """
    Load keyword mapping from keywords.xlsx.
    
    Returns:
        Dict mapping English keywords to Chinese keywords.
        Example: {"Amazon": "亚马逊", "Paladin": "圣骑", ...}
    
    Note:
        This function strictly uses the exact text from Excel.
        No auto-completion, no modification, no inference.
    """
    import sys
    import os
    
    if excel_path is None:
        excel_path = get_keywords_file_path()
    
    mapping = {}
    
    # Try alternative path in _MEIPASS for frozen app
    if not os.path.exists(excel_path) and hasattr(sys, '_MEIPASS'):
        alt_path = os.path.join(sys._MEIPASS, 'keywords.xlsx')
        if os.path.exists(alt_path):
            excel_path = alt_path
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        for row in range(2, ws.max_row + 1):
            en = ws.cell(row=row, column=2).value
            cn = ws.cell(row=row, column=3).value
            if en and cn:
                mapping[str(en)] = str(cn)
        
        return mapping
    
    except Exception as e:
        print(f"Warning: Failed to load keywords from Excel: {e}")
        return {}


# =============================================================================
# Filter Words Configuration (V3.5+)
# =============================================================================

def get_filter_words_file_path() -> str:
    """Get the path to Filter Word Mapping Table.xlsx file."""
    if hasattr(sys, '_MEIPASS'):
        # Running as bundled exe
        return os.path.join(sys._MEIPASS, 'Filter Word Mapping Table.xlsx')
    # Running in normal Python environment
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'Filter Word Mapping Table.xlsx')


def load_filter_words(excel_path: str = None) -> List[str]:
    """
    Load all filter words from Filter Word Mapping Table.xlsx.
    
    Returns:
        List of filter words (column B from Excel).
        Example: ["Faster cast rate", "Faster Hit Recovery", "strength", ...]
    """
    import sys
    import os
    
    if excel_path is None:
        excel_path = get_filter_words_file_path()
    
    filter_words = []
    
    # Try alternative path in _MEIPASS for frozen app
    if not os.path.exists(excel_path) and hasattr(sys, '_MEIPASS'):
        alt_path = os.path.join(sys._MEIPASS, 'Filter Word Mapping Table.xlsx')
        if os.path.exists(alt_path):
            excel_path = alt_path
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # Read column B (filter word) from row 2 onwards
        for row in range(2, ws.max_row + 1):
            filter_word = ws.cell(row=row, column=2).value
            if filter_word and str(filter_word).strip():
                filter_words.append(str(filter_word).strip())
        
        return filter_words
    
    except Exception as e:
        print(f"Warning: Failed to load filter words from Excel: {e}")
        return []
